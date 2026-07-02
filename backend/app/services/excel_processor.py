from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.core.config import DATA_DIR

UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
ALLOWED_EXTENSIONS = {".xlsx"}


@dataclass(frozen=True)
class UploadedExcel:
    file_id: str
    filename: str
    path: Path


def ensure_directories() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _safe_stem(filename: str) -> str:
    stem = Path(filename).stem.strip()
    return "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in stem) or "excel"


def _save_bytes(path: Path, data: bytes) -> None:
    path.write_bytes(data)


def upload_workbook(filename: str, data: bytes) -> UploadedExcel:
    ensure_directories()
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Only .xlsx files are supported")

    file_id = uuid4().hex
    stored_name = f"{file_id}_{_safe_stem(filename)}{suffix}"
    path = UPLOAD_DIR / stored_name
    _save_bytes(path, data)
    return UploadedExcel(file_id=file_id, filename=filename, path=path)


def _load_dataframe(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name or 0, engine="openpyxl")


def read_workbook_metadata(path: Path, sheet_name: str | None = None) -> tuple[int, list[str]]:
    frame = _load_dataframe(path, sheet_name)
    return len(frame), [str(column) for column in frame.columns]


def split_workbook(path: Path, column: str, sheet_name: str | None = None, keep_blank_values: bool = True) -> tuple[list[Path], list[str]]:
    ensure_directories()
    frame = _load_dataframe(path, sheet_name)
    if column not in frame.columns:
        raise KeyError(f"Column not found: {column}")

    working = frame.copy()
    if not keep_blank_values:
        working = working[working[column].notna()]

    output_paths: list[Path] = []
    labels: list[str] = []
    for value, group in working.groupby(column, dropna=not keep_blank_values):
        label = "blank" if pd.isna(value) else str(value)
        safe_label = _safe_stem(label)
        output_path = OUTPUT_DIR / f"{_safe_stem(path.stem)}_{safe_label}.xlsx"
        group.to_excel(output_path, index=False, engine="openpyxl")
        output_paths.append(output_path)
        labels.append(label)

    return output_paths, labels


def merge_workbooks(paths: list[Path], sheet_name: str | None = None, deduplicate: bool = False) -> Path:
    ensure_directories()
    frames = [_load_dataframe(path, sheet_name) for path in paths]
    merged = pd.concat(frames, ignore_index=True)
    if deduplicate:
        merged = merged.drop_duplicates()

    output_path = OUTPUT_DIR / f"merged_{uuid4().hex}.xlsx"
    merged.to_excel(output_path, index=False, engine="openpyxl")
    return output_path


def zip_files(paths: list[Path], zip_name: str) -> Path:
    ensure_directories()
    archive_path = OUTPUT_DIR / f"{_safe_stem(zip_name)}.zip"
    with ZipFile(archive_path, "w", compression=ZIP_DEFLATED) as archive:
        for path in paths:
            archive.write(path, arcname=path.name)
    return archive_path


def inspect_workbook_bytes(data: bytes) -> tuple[str, int]:
    workbook = load_workbook(BytesIO(data))
    sheet = workbook.active
    return str(sheet["A1"].value), int(sheet["B1"].value)

