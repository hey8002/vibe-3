from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook


def build_sample_workbook() -> bytes:
    """Return a small XLSX workbook as bytes for smoke testing or export examples."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sample"
    sheet["A1"] = "label"
    sheet["B1"] = "value"
    sheet["A2"] = "hello"
    sheet["B2"] = 123

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def verify_workbook(data: bytes) -> tuple[str, int]:
    """Read workbook bytes back and return a simple sanity-check summary."""
    workbook = load_workbook(BytesIO(data))
    sheet = workbook.active
    return str(sheet["A2"].value), int(sheet["B2"].value)
